import pandas as pd
from openpyxl.styles import PatternFill



def rgb_to_hex(rgb):
    return '%02x%02x%02x' % rgb


def darken(hex_code, shade):
    shade = shade/10
    #add something to convert shade number ie 9 actually equals 10% darker
    RGB = tuple(int(hex_code[i:i + 2], 16) for i in (0, 2, 4))
    r = RGB[0]
    g = RGB[1]
    b = RGB[2]

    # darken by 10%
    darken_R = int(round(r * shade))
    darken_G = int(round(g * shade))
    darken_B = int(round(b * shade))

    rgbTuple = (darken_R, darken_G, darken_B) #tuple of RGB values to convert to Hex

    return rgb_to_hex(rgbTuple)


def lighten(hex_code, shade):
    shade = shade/10
    RGB = tuple(int(hex_code[i:i + 2], 16) for i in (0, 2, 4))
    r = RGB[0]
    g = RGB[1]
    b = RGB[2]

    # lighten by 10%
    lighten_R = int(round(r + ((255 - r) * shade), 0))
    lighten_G = int(round(g + ((255 - g) * shade), 0))
    lighten_B = int(round(b + ((255 - b) * shade), 0))

    rgbTuple = (lighten_R, lighten_G, lighten_B) #tuple of RGB values to convert to Hex

    return rgb_to_hex(rgbTuple)


def create_lighten_df(colorList):
    empty_df_list = []
    df = pd.DataFrame(columns=['Original Color Hex', 'Change %', 'Color Hex', 'Direction', 'Color'])
    for item in colorList:
        for num in range(11):
            original_color = "#" + item
            change_percent = str(num) + '0%'
            color_hex = "#" + lighten(hex_code=item, shade=num)
            direction = 'Lighter'
            df = pd.concat([df,
                            pd.DataFrame.from_dict({'Original Color Hex': [original_color],
                                                    'Change %': [change_percent],
                                                    'Color Hex': [color_hex],
                                                    'Direction': [direction],
                                                     'Color' : ""})],
                           ignore_index=True)
    return df


def create_darken_df(colorList):
    empty_df_list = []
    df = pd.DataFrame(columns=['Original Color Hex', 'Change %', 'Color Hex', 'Direction', 'Color'])
    for item in colorList:
        for num in range(11):
            original_color = "#" + item
            change_percent = str(num) + '0%'
            color_hex = "#" + darken(hex_code=item, shade=num)
            direction = 'Darken'
            df = pd.concat([df,
                            pd.DataFrame.from_dict({'Original Color Hex': [original_color],
                                                    'Change %': [change_percent],
                                                    'Color Hex': [color_hex],
                                                    'Direction': [direction],
                                                     'Color' : ""})],
                           ignore_index=True)
        print(df)

    percDict = {
        '100%': "00%",
        '90%': "10%",
        '80%': "20%",
        '70%': "30%",
        '60%': "40%",
        '50%': "50%",
        '40%': "60%",
        '30%': "70%",
        '20%': "80%",
        '10%': "90%",
        '00%': "100%"
    }
    df['Change %'] = df['Change %'].apply(lambda x: percDict[x])
    df['Original Color Hex'] = df['Original Color Hex'].astype('category')
    df = df.sort_index(ascending=False).reset_index(drop=True)
    df['Sorting'] = df.index
    sorter = ['#' + str(x) for x in colorList]
    df['Original Color Hex'] = df['Original Color Hex'].cat.set_categories(sorter)
    df = df.sort_values(by=['Original Color Hex', 'Sorting']).drop(columns='Sorting')

    return df


###############start of work###################

colorList = ['CC0000', #red
             '000000', #black
             '4d4d4f', #dark gray
             '969697', #medium gray
             'dddddd', # light gray
             'f3f3f3', # ultra light gray
             'f58025', # ada compliant orange
             'fdb913', # ada compliant yellow
             '97ca3e', # ada compliant green
             '479cd6', # ada compliant blue
             '1d3c6d', # ada compliant navy
             '751c59' # ada compliant purple
             ]

lighten_df = create_lighten_df(colorList)
darken_df = create_darken_df(colorList)


with pd.ExcelWriter("ColorShadeReferenceWorkbook.xlsx", engine="openpyxl") as writer:
    # create the "Lighten" worksheet
    sheet_name = "Lighten"
    # Export DataFrame content
    lighten_df.to_excel(writer, sheet_name=sheet_name, index=False)
    # Set column width dimensions
    sheet = writer.sheets[sheet_name]# open sheet
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 9
    # Set background colors depending on cell values
    for cell, in sheet[f'E2:E{len(lighten_df) + 1}']: # Skip header row, process as many rows as there are DataFrames
        value = lighten_df["Color Hex"].iloc[cell.row - 2] #set color value to row['Color Hex'] value
        value = value.strip("#")
        cell.fill = PatternFill(start_color=value, end_color=value, fill_type='solid')

    #create the "Darken" worksheet
    sheet_name = "Darken"
    # Export DataFrame content
    darken_df.to_excel(writer, sheet_name=sheet_name, index=False)
    sheet = writer.sheets[sheet_name]
    # Set column width dimensions
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 9
    # Set background colors depending on cell values
    for cell, in sheet[f'E2:E{len(darken_df) + 1}']: # Skip header row, process as many rows as there are DataFrames
        value = darken_df["Color Hex"].iloc[cell.row - 2] #set color value to row['Color Hex'] value
        value = value.strip("#") #strip #
        cell.fill = PatternFill(start_color=value, end_color=value, fill_type='solid') #fill cell colid color